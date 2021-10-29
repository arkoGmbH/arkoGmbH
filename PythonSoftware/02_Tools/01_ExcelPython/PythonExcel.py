# Write 'Hello, World' to standard output.
import sys

print("Hello Pythones!...")

import openpyxl
path='Temp.xlsm'

wb = openpyxl.load_workbook(path)
ws = wb.get_sheet_by_name('MainGUI')
ws['B4']='koadfk'
cell = ws.cell(row = 4, column = 2)
print(cell.value)
wb.save(filename = 'Temp.xlsm')
