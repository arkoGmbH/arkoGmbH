# Write 'Hello, World' to standard output.
import sys
import tkinter as tk
from openpyxl import Workbook
print("Hello Pythones!...")
# Show the currently used interpreter
#print(sys.executable)
#print(sys.path)


root= tk.Tk()

canvas1 = tk.Canvas(root, width = 300, height = 300)
canvas1.pack()

# Instantiate workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['15'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")



def hello ():  
    label1 = tk.Label(root, text= 'Hello World!', fg='green', font=('helvetica', 12, 'bold'))
    canvas1.create_window(150, 200, window=label1)
    
button1 = tk.Button(text='----',command=hello, bg='brown',fg='white')
canvas1.create_window(150, 150, window=button1)

root.mainloop()